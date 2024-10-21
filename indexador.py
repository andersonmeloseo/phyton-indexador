import re
import sys
from oauth2client.service_account import ServiceAccountCredentials
from googleapiclient.discovery import build
import requests
import xml.etree.ElementTree as ET
import datetime
import csv
from openpyxl import Workbook

# Configuração inicial
SCOPES = ["https://www.googleapis.com/auth/indexing", "https://www.googleapis.com/auth/webmasters"]
JSON_KEY_FILE = "personal.json"

# Função para validar a URL do domínio
def validate_domain_input(domain):
    domain = domain.strip().lower()
    domain = domain.replace("httos://", "https://").replace("htto://", "http://")
    
    if not domain.startswith("http://") and not domain.startswith("https://"):
        raise ValueError("O domínio precisa começar com 'http://' ou 'https://'.")
    
    if len(domain.split('.')) < 2:
        raise ValueError("Domínio inválido. Verifique a URL e tente novamente.")
    
    return domain

# Solicitar a URL do domínio a ser analisado
try:
    DOMAIN = input("Digite a URL do domínio que deseja analisar (exemplo: https://cbmadvs.com.br/): ").strip()
    DOMAIN = validate_domain_input(DOMAIN)
except ValueError as e:
    print(e)
    sys.exit(1)

# Autenticação com as credenciais
credentials = ServiceAccountCredentials.from_json_keyfile_name(JSON_KEY_FILE, scopes=SCOPES)
search_service = build('searchconsole', 'v1', credentials=credentials)
indexing_service = build('indexing', 'v3', credentials=credentials)

# Função para obter URLs de um sitemap
def get_urls_from_sitemap(sitemap_url):
    try:
        response = requests.get(sitemap_url)
        response.raise_for_status()
        urls = []
        root = ET.fromstring(response.content)

        # Verifica se é um índice de sitemaps (referenciando outros sitemaps)
        for sitemap in root.findall(".//{http://www.sitemaps.org/schemas/sitemap/0.9}sitemap"):
            loc = sitemap.find("{http://www.sitemaps.org/schemas/sitemap/0.9}loc").text
            urls.extend(get_urls_from_sitemap(loc))
        
        # Caso seja o sitemap com as URLs diretamente
        for url in root.findall(".//{http://www.sitemaps.org/schemas/sitemap/0.9}loc"):
            urls.append(url.text)

        print(f"Encontradas {len(urls)} URLs no sitemap: {sitemap_url}")
        return urls
    except requests.exceptions.RequestException as e:
        print(f"Erro ao acessar o sitemap {sitemap_url}: {e}")
        return []

# Função para obter todas as URLs dos sitemaps listados no Search Console
def get_all_urls_from_sitemaps(service, site_url):
    try:
        sitemap_list = service.sitemaps().list(siteUrl=site_url).execute()
    except Exception as e:
        print(f"Erro ao acessar a API do Search Console: {e}")
        return []
    
    all_urls = []
    for sitemap in sitemap_list.get('sitemap', []):
        sitemap_url = sitemap['path']
        print(f"Processando sitemap: {sitemap_url}")
        all_urls.extend(get_urls_from_sitemap(sitemap_url))
    
    return list(set(all_urls))

# Função para obter o status de indexação e dados de desempenho das URLs com paginação
def get_index_status(service, site_url):
    all_rows = []
    start_row = 0
    page_size = 1000
    while True:
        try:
            request = service.searchanalytics().query(
                siteUrl=site_url,
                body={
                    "startDate": "2023-01-01",
                    "endDate": datetime.date.today().strftime('%Y-%m-%d'),
                    "dimensions": ["page"],  # Captura URLs
                    "startRow": start_row,
                    "rowLimit": page_size
                }
            )
            response = request.execute()
            rows = response.get('rows', [])
            if not rows:
                break
            all_rows.extend(rows)
            start_row += len(rows)
            print(f"Obtidos {len(rows)} URLs.")
            if len(rows) < page_size:
                break
        except Exception as e:
            print(f"Erro ao obter status de indexação: {e}")
            break
    return all_rows

# Função para obter palavras-chave
def get_keywords(service, site_url):
    all_rows = []
    start_row = 0
    page_size = 1000
    while True:
        try:
            request = service.searchanalytics().query(
                siteUrl=site_url,
                body={
                    "startDate": "2023-01-01",
                    "endDate": datetime.date.today().strftime('%Y-%m-%d'),
                    "dimensions": ["query"],  # Captura palavras-chave
                    "startRow": start_row,
                    "rowLimit": page_size
                }
            )
            response = request.execute()
            rows = response.get('rows', [])
            if not rows:
                break
            all_rows.extend(rows)
            start_row += len(rows)
            print(f"Obtidos {len(rows)} palavras-chave.")
            if len(rows) < page_size:
                break
        except Exception as e:
            print(f"Erro ao obter palavras-chave: {e}")
            break
    return all_rows

# Função para obter comparativo dos últimos 30 dias com os 30 dias anteriores
def get_performance_comparison(service, site_url):
    end_date = datetime.date.today()
    start_date_30_days = end_date - datetime.timedelta(days=30)
    start_date_60_days = start_date_30_days - datetime.timedelta(days=30)

    # Dados dos últimos 30 dias
    request_30_days = service.searchanalytics().query(
        siteUrl=site_url,
        body={
            "startDate": start_date_30_days.strftime('%Y-%m-%d'),
            "endDate": end_date.strftime('%Y-%m-%d'),
            "dimensions": ["page"]
        }
    ).execute()

    # Dados dos 30 dias anteriores
    request_30_days_before = service.searchanalytics().query(
        siteUrl=site_url,
        body={
            "startDate": start_date_60_days.strftime('%Y-%m-%d'),
            "endDate": start_date_30_days.strftime('%Y-%m-%d'),
            "dimensions": ["page"]
        }
    ).execute()

    return request_30_days.get('rows', []), request_30_days_before.get('rows', [])

# Função para enviar uma URL para a fila de indexação com tratamento de erros
def send_url_to_indexing(service, url):
    body = {
        "url": url,
        "type": "URL_UPDATED"
    }
    try:
        response = service.urlNotifications().publish(body=body).execute()
        return response, "Enviado para indexação"
    except Exception as e:
        print(f"Erro ao enviar a URL: {url}")
        print(f"Detalhes do erro: {e}")
        return str(e), "Erro ao enviar para indexação"

# Função para perguntar quantas URLs enviar para indexação
def ask_how_many_urls_to_index(non_indexed_urls):
    max_limit = min(200, len(non_indexed_urls))  # Limita a 200 URLs por dia ou menos
    while True:
        try:
            num_urls = int(input(f"Quantas URLs deseja enviar para indexação? (Máximo de {max_limit}, ou 0 para não enviar): "))
            if 0 <= num_urls <= max_limit:
                return num_urls
            else:
                print(f"Por favor, insira um número entre 0 e {max_limit}.")
        except ValueError:
            print("Entrada inválida. Por favor, insira um número válido.")

# Função para gerar relatório em Excel
def create_excel_report(url_rows, keyword_rows, all_urls, indexed_urls, non_indexed_urls, sent_for_indexing, performance_30_days, performance_30_days_before):
    wb = Workbook()
    
    # Aba 1: Relatório de Indexação
    ws_index = wb.active
    ws_index.title = "Relatório de Indexação"
    ws_index.append(["URL", "Status", "Ação", "Impressões"])

    for url in all_urls:
        status = "Indexada" if url in indexed_urls else "Rastreadas, mas não indexadas"
        impressions = indexed_urls.get(url, 0)  # Obtém impressões se a URL estiver indexada
        action = next((item[1] for item in sent_for_indexing if item[0] == url), "")
        ws_index.append([url, status, action, impressions])

    # Aba 2: Relatório de Palavras-chave
    ws_keywords = wb.create_sheet("Palavras-chave")
    ws_keywords.append(["Palavra-chave", "Cliques", "Impressões", "CTR"])
    for row in keyword_rows:
        keywords = row['keys'][0]
        clicks = row.get('clicks', 0)
        impressions = row.get('impressions', 0)
        ctr = row.get('ctr', 0)
        ws_keywords.append([keywords, clicks, impressions, ctr])

    # Aba 3: Comparativo dos últimos 30 dias e 30 dias anteriores
    ws_performance = wb.create_sheet("Comparativo Desempenho")
    ws_performance.append([
        "URL", "Cliques (Últimos 30 dias)", "Impressões (Últimos 30 dias)", 
        "Posição Média (Últimos 30 dias)", "Cliques (30 dias anteriores)", 
        "Impressões (30 dias anteriores)", "Posição Média (30 dias anteriores)", 
        "Diferença Cliques (%)", "Diferença Impressões (%)", "Diferença Posição (%)"
    ])

    # Comparar os dados dos dois períodos
    clicks_30_days = {row['keys'][0]: row.get('clicks', 0) for row in performance_30_days}
    impressions_30_days = {row['keys'][0]: row.get('impressions', 0) for row in performance_30_days}
    position_30_days = {row['keys'][0]: row.get('position', 0) for row in performance_30_days}

    clicks_30_days_before = {row['keys'][0]: row.get('clicks', 0) for row in performance_30_days_before}
    impressions_30_days_before = {row['keys'][0]: row.get('impressions', 0) for row in performance_30_days_before}
    position_30_days_before = {row['keys'][0]: row.get('position', 0) for row in performance_30_days_before}

    # Calcula as diferenças percentuais e adiciona ao relatório
    for url in set(clicks_30_days.keys()).union(clicks_30_days_before.keys()):
        clicks_now = clicks_30_days.get(url, 0)
        impressions_now = impressions_30_days.get(url, 0)
        position_now = position_30_days.get(url, 0)

        clicks_before = clicks_30_days_before.get(url, 0)
        impressions_before = impressions_30_days_before.get(url, 0)
        position_before = position_30_days_before.get(url, 0)

        # Diferenças percentuais (evita divisão por zero)
        change_clicks = ((clicks_now - clicks_before) / clicks_before * 100) if clicks_before > 0 else 100 if clicks_now > 0 else 0
        change_impressions = ((impressions_now - impressions_before) / impressions_before * 100) if impressions_before > 0 else 100 if impressions_now > 0 else 0
        change_position = ((position_now - position_before) / position_before * 100) if position_before > 0 else 0

        ws_performance.append([
            url, clicks_now, impressions_now, position_now,
            clicks_before, impressions_before, position_before,
            change_clicks, change_impressions, change_position
        ])

    # Aba 4: Lista de URLs com 'site:' e status
    ws_urls = wb.create_sheet("URLs para Inspeção")
    ws_urls.append(["URL", "Status"])
    for url in all_urls:
        status = "Indexada" if url in indexed_urls else "Rastreadas, mas não indexadas"
        ws_urls.append([f"site:{url}", status])

    # Salvar o relatório
    now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    domain_name = DOMAIN.replace("https://", "").replace("http://", "").replace("/", "")
    filename = f"relatorio_indexacao_{domain_name}_{now}.xlsx"
    wb.save(filename)
    print(f"Relatório gerado: {filename}")

# Função para registrar URLs enviadas para indexação
def log_indexing_results(sent_for_indexing):
    with open('indexing_log.csv', mode='a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['URL', 'Status', 'Timestamp'])
        for url, status in sent_for_indexing:
            writer.writerow([url, status, datetime.datetime.now()])
    print("Log de indexação atualizado.")

# Fluxo principal de execução
def main():
    all_urls = get_all_urls_from_sitemaps(search_service, DOMAIN)
    if len(all_urls) == 0:
        print("Nenhuma URL encontrada no sitemap.")
        return

    url_rows = get_index_status(search_service, DOMAIN)  # Dados das URLs
    keyword_rows = get_keywords(search_service, DOMAIN)  # Dados das palavras-chave

    indexed_urls = {row['keys'][0]: row.get('impressions', 0) for row in url_rows}
    non_indexed_urls = set(all_urls) - indexed_urls.keys()

    # Obter dados de desempenho dos últimos 30 dias e dos 30 dias anteriores
    performance_30_days, performance_30_days_before = get_performance_comparison(search_service, DOMAIN)

    # Perguntar quantas URLs o usuário quer enviar para indexação
    num_urls_to_index = ask_how_many_urls_to_index(non_indexed_urls)

    sent_for_indexing = []
    if num_urls_to_index > 0:
        request_count = 0
        for url in non_indexed_urls:
            if request_count >= num_urls_to_index:
                break
            response, status = send_url_to_indexing(indexing_service, url)
            sent_for_indexing.append((url, status))
            print(f"URL: {url}, Status: {status}")
            request_count += 1

    log_indexing_results(sent_for_indexing)

    # Gerar o relatório final
    create_excel_report(url_rows, keyword_rows, all_urls, indexed_urls, non_indexed_urls, sent_for_indexing, performance_30_days, performance_30_days_before)

if __name__ == "__main__":
    main()
